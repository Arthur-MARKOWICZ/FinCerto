"""
Serviço de geração de relatórios financeiros.

Implementa o padrão Strategy com geradores intercambiáveis (Excel/PDF).
Converte dados do banco em DataFrames do Pandas para manipulação e
gera arquivos formatados para download.
"""

import logging
from abc import ABC, abstractmethod
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .report_repository import ReportRepository

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


class ReportGenerator(ABC):
    """
    Classe base abstrata para geradores de relatório.

    Define a interface comum (generate) e métodos de validação
    compartilhados entre os formatos Excel e PDF.
    """

    @abstractmethod
    def generate(self, data: List[Any], **kwargs) -> Tuple[BytesIO, str, str]:
        """
        Gera o relatório a partir de uma lista de dados.

        Args:
            data: Lista de registros (dicts ou Row objects) do banco.
            **kwargs: Parâmetros adicionais específicos do gerador.

        Returns:
            Tupla com (buffer do arquivo, mime-type, nome do arquivo).
        """
        pass

    def validate_data(self, data: List[Any]) -> None:
        """
        Valida os dados de entrada antes da geração.

        Args:
            data: Dados a serem validados.

        Raises:
            ValueError: Se data for None ou vazio.
            TypeError: Se data não for lista ou tupla.
        """
        if data is None:
            raise ValueError("Dados não podem ser None")
        if not isinstance(data, (list, tuple)):
            raise TypeError(
                f"Os dados devem ser uma lista ou tupla, "
                f"recebido: {type(data).__name__}"
            )
        if len(data) == 0:
            raise ValueError("Nenhum dado fornecido para gerar o relatório")

        logger.debug(f"Validação de dados OK. Total de registros: {len(data)}")

    def _safe_get_value(self, obj: Any, field: str, default: str = "") -> str:
        """
        Extrai um valor de forma segura de um objeto (dict ou objeto com atributos).

        Args:
            obj: Objeto fonte (dict, Row, dataclass, etc.).
            field: Nome do campo a extrair.
            default: Valor padrão se o campo não existir.

        Returns:
            Valor convertido para string.
        """
        try:
            if isinstance(obj, dict):
                return str(obj.get(field, default))
            elif hasattr(obj, field):
                return str(getattr(obj, field, default))
            return default
        except Exception as e:
            logger.warning(f"Erro ao extrair campo '{field}': {e}")
            return default


class ExcelGenerator(ReportGenerator):
    """
    Gera relatórios em formato Excel (.xlsx) com formatação visual.

    Utiliza Pandas + openpyxl para criar planilhas estilizadas
    com cabeçalho colorido e colunas dimensionadas.
    """

    def __init__(
        self,
        columns: List[Dict[str, str]],
        sheet_name: str = "Relatório",
        title: str = None,
    ) -> None:
        self.columns = columns
        self.sheet_name = sheet_name
        self.title = title
        self.logger = logging.getLogger(self.__class__.__name__)

    def _format_data(self, data: List[Any]) -> List[Dict[str, Any]]:
        """
        Converte dados brutos em lista de dicionários normalizados.

        Garante que cada registro contenha ao menos os campos mínimos
        (data, descricao, categoria, valor, tipo).

        Args:
            data: Lista de registros do banco de dados.

        Returns:
            Lista de dicionários com chaves padronizadas.
        """
        formatted_data: List[Dict[str, Any]] = []

        for i, row in enumerate(data):
            try:
                row_dict = self._row_to_dict(row)

                if not row_dict:
                    self.logger.warning(
                        f"Linha {i} veio vazia ou não pôde ser convertida: {row}"
                    )
                    continue

                # Garantir campos mínimos
                row_dict.setdefault("data", "")
                row_dict.setdefault("descricao", "")
                row_dict.setdefault("categoria", "")
                row_dict.setdefault("valor", 0.0)
                row_dict.setdefault("tipo", "")

                formatted_data.append(row_dict)
            except Exception as e:
                self.logger.error(
                    f"Erro ao formatar linha {i}: {e}, dados: {row}"
                )
                formatted_data.append(
                    {
                        "data": "",
                        "descricao": "",
                        "categoria": "",
                        "valor": 0.0,
                        "tipo": "",
                    }
                )

        if not formatted_data:
            self.logger.warning("Nenhum dado foi formatado corretamente")

        return formatted_data

    def _row_to_dict(self, row: Any) -> Dict[str, Any]:
        """
        Converte um registro (Row, tuple, dict) em dicionário.

        Suporta múltiplos formatos de entrada:
        - dict: retorna diretamente
        - tuple/list/SQLAlchemy Row: mapeia por índice posicional
        - objeto com __dict__: extrai atributos públicos

        Args:
            row: Registro a converter.

        Returns:
            Dicionário com os campos extraídos, ou dict vazio em caso de falha.
        """
        self.logger.debug(f"Tentando converter: {type(row)} - {row}")

        if isinstance(row, dict):
            return row
        elif isinstance(row, (tuple, list)) or "sqlalchemy.engine.row.Row" in str(
            type(row)
        ):
            try:
                if len(row) >= 6:
                    result = {
                        "data": row[0] if row[0] is not None else "",
                        "descricao": row[1] if row[1] is not None else "",
                        "categoria": row[2] if row[2] is not None else "",
                        "valor": float(row[4]) if row[4] is not None else 0.0,
                        "tipo": str(row[5]) if row[5] is not None else "",
                    }
                    self.logger.debug(f"Conversão bem sucedida: {result}")
                    return result
                else:
                    self.logger.warning(
                        f"Registro com tamanho insuficiente: {row} "
                        f"(esperado >=6, got {len(row)})"
                    )
            except (IndexError, ValueError, TypeError) as e:
                self.logger.error(f"Erro ao converter registro {row}: {e}")
        elif hasattr(row, "__dict__"):
            try:
                return {
                    k: v for k, v in row.__dict__.items() if not k.startswith("_")
                }
            except Exception as e:
                self.logger.error(f"Erro ao converter objeto com __dict__: {e}")
        else:
            self.logger.warning(f"Tipo não suportado: {type(row)} - {row}")

        return {}

    def generate(self, data: List[Any], **kwargs) -> Tuple[BytesIO, str, str]:
        """
        Gera o relatório em formato Excel (.xlsx).

        Converte os dados em DataFrame do Pandas, aplica formatação visual
        (cabeçalho com fundo azul e texto branco) e retorna o buffer pronto.

        Args:
            data: Lista de registros a incluir no relatório.
            **kwargs: Parâmetros adicionais (não utilizados).

        Returns:
            Tupla com (BytesIO, mime-type, nome do arquivo).

        Raises:
            ValueError: Se os dados forem inválidos ou vazios.
        """
        self.validate_data(data)

        formatted_data = self._format_data(data)

        if not formatted_data:
            self.logger.warning(
                "Dados formatados estão vazios — criando DataFrame com colunas padrão"
            )
            df = pd.DataFrame(columns=[col["field"] for col in self.columns])
        else:
            df = pd.DataFrame(formatted_data)

        available_columns = [
            col["field"] for col in self.columns if col["field"] in df.columns
        ]
        missing_columns = [
            col["field"] for col in self.columns if col["field"] not in df.columns
        ]

        if missing_columns:
            self.logger.warning(f"Colunas ausentes no DataFrame: {missing_columns}")
            for col in missing_columns:
                df[col] = ""

        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(
                writer,
                index=False,
                sheet_name=self.sheet_name,
                columns=available_columns,
            )

            # Aplicar estilo no cabeçalho
            try:
                from openpyxl.styles import Alignment, Font, PatternFill

                ws = writer.sheets[self.sheet_name]
                fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                font = Font(bold=True, color="FFFFFF")
                alignment = Alignment(horizontal="center", vertical="center")

                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = alignment
            except Exception:
                pass

        buffer.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relatorio_{timestamp}.xlsx"

        self.logger.info(f"Relatório Excel gerado: {filename}")
        return (
            buffer,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename,
        )


class PDFGenerator(ReportGenerator):
    """
    Gera relatórios em formato PDF com estilo profissional.

    Utiliza ReportLab para criar documentos com título, tabela formatada,
    rodapé e data de emissão.
    """

    def __init__(self, title: str, columns: List[Dict[str, str]]) -> None:
        self.title = title
        self.columns = columns
        self.logger = logging.getLogger(self.__class__.__name__)

    def _build_table_data(self, data: List[Any]) -> List[List[str]]:
        """
        Monta os dados da tabela PDF a partir dos registros.

        Formata valores monetários no padrão brasileiro (R$ 1.234,56)
        e datas no formato dd/mm/yyyy.

        Args:
            data: Lista de registros (dicts ou objetos).

        Returns:
            Lista de listas de strings (cabeçalho + linhas de dados).
        """
        table_data: List[List[str]] = []

        # Cabeçalho
        header = [col.get("header", col["field"]) for col in self.columns]
        table_data.append(header)

        for row in data:
            table_row: List[str] = []
            for col in self.columns:
                field = col["field"]

                if isinstance(row, dict):
                    value = row.get(field, "")
                else:
                    value = getattr(row, field, "")

                # Formatação de tipos
                if isinstance(value, datetime):
                    value = value.strftime("%d/%m/%Y")
                elif isinstance(value, (int, float)):
                    if col.get("format") == "currency":
                        value = (
                            f"R$ {value:,.2f}"
                            .replace(".", "|")
                            .replace(",", ".")
                            .replace("|", ",")
                        )
                    else:
                        value = (
                            f"{value:,.2f}"
                            .replace(".", "|")
                            .replace(",", ".")
                            .replace("|", ",")
                        )

                # Truncar textos longos
                value_str = str(value)
                if len(value_str) > 40:
                    value_str = value_str[:37] + "..."

                table_row.append(value_str)

            table_data.append(table_row)

        return table_data

    def _create_styled_table(self, table_data: List[List[str]]) -> Table:
        """
        Cria uma tabela ReportLab com estilo profissional.

        Aplica cabeçalho escuro, linhas alternadas, bordas sutis
        e alinhamento adequado para cada tipo de coluna.

        Args:
            table_data: Dados da tabela (cabeçalho + linhas).

        Returns:
            Objeto Table do ReportLab estilizado.
        """
        table = Table(
            table_data, repeatRows=1, colWidths=[100] * len(self.columns)
        )

        header_bg = HexColor("#2C3E50")
        header_text = colors.whitesmoke
        row_bg_light = HexColor("#ECF0F1")
        row_bg_dark = HexColor("#FFFFFF")
        border_color = HexColor("#BDC3C7")

        style_commands = [
            # Cabeçalho
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), header_text),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 14),
            ("TOPPADDING", (0, 0), (-1, 0), 14),
            # Dados
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [row_bg_dark, row_bg_light]),
            ("TEXTCOLOR", (0, 1), (-1, -1), HexColor("#2C3E50")),
            ("ALIGN", (0, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            # Bordas e padding
            ("GRID", (0, 0), (-1, -1), 1, border_color),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 1), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 10),
        ]

        table.setStyle(TableStyle(style_commands))
        return table

    def generate(self, data: List[Any], **kwargs) -> Tuple[BytesIO, str, str]:
        """
        Gera o relatório completo em formato PDF.

        Cria documento A4 com título centralizado, data de emissão,
        tabela de dados estilizada e rodapé institucional.

        Args:
            data: Lista de registros a incluir no relatório.
            **kwargs: Parâmetros adicionais (não utilizados).

        Returns:
            Tupla com (BytesIO, mime-type, nome do arquivo).

        Raises:
            ValueError: Se os dados forem inválidos ou vazios.
        """
        self.validate_data(data)

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
        )

        elements = []

        # Título
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=getSampleStyleSheet()["Heading1"],
            fontSize=18,
            textColor=HexColor("#2C3E50"),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
        )
        elements.append(Paragraph(self.title, title_style))

        # Data de emissão
        date_style = ParagraphStyle(
            "DateStyle",
            parent=getSampleStyleSheet()["Normal"],
            fontSize=9,
            textColor=HexColor("#7F8C8D"),
            alignment=TA_RIGHT,
            spaceAfter=12,
        )
        elements.append(
            Paragraph(
                f"Emitido em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}",
                date_style,
            )
        )

        elements.append(Spacer(1, 0.2 * inch))

        # Tabela de dados
        table_data = self._build_table_data(data)
        if table_data and len(table_data) > 1:
            table = self._create_styled_table(table_data)
            elements.append(table)
        else:
            elements.append(
                Paragraph(
                    "Nenhum dado disponível para o relatório.",
                    getSampleStyleSheet()["Normal"],
                )
            )

        # Rodapé
        elements.append(Spacer(1, 0.3 * inch))
        footer_style = ParagraphStyle(
            "Footer",
            parent=getSampleStyleSheet()["Normal"],
            fontSize=8,
            textColor=HexColor("#95A5A6"),
            alignment=TA_CENTER,
        )
        elements.append(
            Paragraph(
                "Sistema FinanCerto - Relatório Gerado Automaticamente",
                footer_style,
            )
        )

        doc.build(elements)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relatorio_{timestamp}.pdf"

        self.logger.info(f"Relatório PDF gerado com sucesso: {filename}")
        return buffer, "application/pdf", filename


class ReportService:
    """
    Serviço principal de geração de relatórios.

    Orquestra a busca de dados via ReportRepository e a geração
    de arquivos via geradores (Excel/PDF) usando o padrão Strategy.
    """

    REPORTS: Dict[str, Dict[str, Any]] = {
        "transacoes_por_categoria": {
            "title": "Relatório de Transações por Categoria",
            "columns": [
                {"field": "data", "header": "Data", "width": 80, "format": "date"},
                {
                    "field": "descricao",
                    "header": "Descrição",
                    "width": 150,
                    "multiline": True,
                },
                {"field": "categoria", "header": "Categoria", "width": 100},
                {
                    "field": "valor",
                    "header": "Valor",
                    "width": 80,
                    "format": "currency",
                },
                {"field": "tipo", "header": "Tipo", "width": 60},
            ],
        },
        "saldo_mensal": {
            "title": "Relatório de Saldo Mensal",
            "columns": [
                {"field": "mes_nome", "header": "Mês/Ano", "width": 120},
                {
                    "field": "receitas",
                    "header": "Receitas (R$)",
                    "width": 100,
                    "format": "currency",
                },
                {
                    "field": "despesas",
                    "header": "Despesas (R$)",
                    "width": 100,
                    "format": "currency",
                },
                {
                    "field": "saldo",
                    "header": "Saldo (R$)",
                    "width": 100,
                    "format": "currency",
                },
            ],
        },
        "transacoes_detalhadas": {
            "title": "Relatório de Transações Detalhadas",
            "columns": [
                {"field": "data_formatada", "header": "Data", "width": 350},
                {
                    "field": "descricao",
                    "header": "Descrição",
                    "width": 350,
                    "multiline": True,
                },
                {"field": "categoria", "header": "Categoria", "width": 350},
                {"field": "conta", "header": "Conta", "width": 350},
                {"field": "tipo_transacao", "header": "Tipo", "width": 350},
                {
                    "field": "valor",
                    "header": "Valor (R$)",
                    "width": 350,
                    "format": "currency",
                },
            ],
        },
    }

    def __init__(self, report_repository: ReportRepository) -> None:
        self.report_repository = report_repository
        self.logger = logging.getLogger(self.__class__.__name__)

    def _validate_report_type(self, report_type: str) -> None:
        """
        Valida se o tipo de relatório solicitado é suportado.

        Args:
            report_type: Identificador do tipo de relatório.

        Raises:
            ValueError: Se o tipo for vazio ou não estiver em REPORTS.
        """
        if not report_type:
            raise ValueError("Tipo de relatório não pode estar vazio")
        if report_type not in self.REPORTS:
            valid_types = ", ".join(self.REPORTS.keys())
            raise ValueError(
                f"Tipo de relatório inválido: '{report_type}'. "
                f"Tipos válidos: {valid_types}"
            )
        self.logger.debug(f"Tipo de relatório validado: {report_type}")

    def _validate_format_type(self, format_type: str) -> str:
        """
        Valida e normaliza o formato de saída do relatório.

        Args:
            format_type: Formato solicitado ('excel' ou 'pdf').

        Returns:
            Formato normalizado em minúsculas.

        Raises:
            ValueError: Se o formato for inválido.
        """
        if not format_type:
            raise ValueError("Tipo de formato não pode estar vazio")

        normalized_format = format_type.lower().strip()
        if normalized_format not in ["excel", "pdf"]:
            raise ValueError(
                f"Formato inválido: '{format_type}'. "
                f"Formatos suportados: excel, pdf"
            )

        self.logger.debug(f"Formato de saída validado: {normalized_format}")
        return normalized_format

    def _get_generator(
        self, report_type: str, format_type: str
    ) -> ReportGenerator:
        """
        Cria o gerador apropriado (Strategy) baseado no tipo e formato.

        Args:
            report_type: Tipo de relatório (chave em REPORTS).
            format_type: Formato de saída ('excel' ou 'pdf').

        Returns:
            Instância de ExcelGenerator ou PDFGenerator.
        """
        self._validate_report_type(report_type)
        normalized_format = self._validate_format_type(format_type)

        report_config = self.REPORTS[report_type]

        if normalized_format == "excel":
            self.logger.info(f"Criando ExcelGenerator para: {report_type}")
            return ExcelGenerator(
                columns=report_config["columns"],
                sheet_name=report_config["title"][:31],
            )
        else:
            self.logger.info(f"Criando PDFGenerator para: {report_type}")
            return PDFGenerator(
                title=report_config["title"],
                columns=report_config["columns"],
            )

    async def gerar_relatorio(
        self,
        report_type: str,
        format_type: str = "excel",
        **filters: Any,
    ) -> Tuple[BytesIO, str, str]:
        """
        Método principal de geração de relatórios.

        Busca dados no repositório, valida, e delega a geração
        ao gerador correto (Excel ou PDF).

        Args:
            report_type: Tipo de relatório
                ('transacoes_por_categoria', 'saldo_mensal', 'transacoes_detalhadas').
            format_type: Formato de saída ('excel' ou 'pdf').
            **filters: Filtros de busca (usuario_id, categoria_id, etc.).

        Returns:
            Tupla com (BytesIO do arquivo, mime-type, nome do arquivo).

        Raises:
            ValueError: Se os dados ou parâmetros forem inválidos.
        """
        start_time = datetime.now()
        self.logger.info(
            f"Iniciando geração de relatório: type={report_type}, "
            f"format={format_type}"
        )

        self._validate_report_type(report_type)
        self._validate_format_type(format_type)

        # Buscar dados conforme o tipo de relatório
        try:
            if report_type == "saldo_mensal":
                data = self.report_repository.buscar_saldo_mensal(
                    usuario_id=filters.get("usuario_id"),
                    ano=filters.get("ano"),
                    conta_id=filters.get("conta_id"),
                )
            elif report_type == "transacoes_detalhadas":
                data = self.report_repository.buscar_transacoes_detalhadas(
                    usuario_id=filters.get("usuario_id"),
                    conta_id=filters.get("conta_id"),
                    categoria_id=filters.get("categoria_id"),
                    data_inicio=filters.get("data_inicio"),
                    data_fim=filters.get("data_fim"),
                    valor_minimo=filters.get("valor_minimo"),
                    valor_maximo=filters.get("valor_maximo"),
                    tipo_transacao=filters.get("tipo_transacao"),
                )
            else:
                data = self.report_repository.buscar_relatorio(**filters)
        except Exception as e:
            self.logger.error(f"Erro ao buscar dados para relatório: {e}")
            raise ValueError(f"Erro ao recuperar dados do relatório: {str(e)}")

        if not data:
            self.logger.warning(
                f"Nenhum dado encontrado para {report_type} "
                f"com filtros: {filters}"
            )
            raise ValueError("Nenhum dado encontrado para os filtros fornecidos")

        self.logger.debug(f"Dados recuperados: {len(data)} registros")

        # Gerar relatório
        generator = self._get_generator(report_type, format_type)
        buffer, mime_type, filename = generator.generate(data)

        elapsed_time = (datetime.now() - start_time).total_seconds()
        self.logger.info(
            f"Relatório gerado com sucesso: {filename} "
            f"({len(data)} registros em {elapsed_time:.2f}s)"
        )

        return buffer, mime_type, filename

    async def gerar_relatorio_por_categoria(
        self,
        usuario_id: int,
        categoria_id: int = None,
        tipo: str = None,
        periodo: str = None,
        formato: str = "excel",
    ) -> Tuple[BytesIO, str, str]:
        """
        Atalho para geração de relatório de transações por categoria.

        Args:
            usuario_id: ID do usuário proprietário.
            categoria_id: ID da categoria para filtrar (opcional).
            tipo: Tipo de transação (opcional).
            periodo: Período temporal (opcional).
            formato: Formato de saída — 'excel' ou 'pdf'.

        Returns:
            Tupla com (BytesIO do arquivo, mime-type, nome do arquivo).
        """
        return await self.gerar_relatorio(
            report_type="transacoes_por_categoria",
            format_type=formato,
            usuario_id=usuario_id,
            categoria_id=categoria_id,
            tipo=tipo,
            periodo=periodo,
        )